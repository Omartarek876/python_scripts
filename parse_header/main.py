
"""
==============================================================================
Script Name   : parse_header_file.py
Author        : Omar Tarek
Date Created  : September 4, 2024
Description   : This script parses a C header file to extract function 
                prototypes, global variable declarations, macros, and 
                include statements. The results are saved in an Excel 
                workbook with separate sheets for each category.
==============================================================================
"""

import openpyxl
import re
import os
from openpyxl.styles import PatternFill


def parse_hf(header_path):
    with open(header_path, "r") as file:
        content = file.read()

    # Regular expressions (excluding preprocessor directives)
    patterns = {
        'function_pattern': re.compile(r'\b[A-Za-z_][A-Za-z0-9_]*\s+\**\b[A-Za-z_][A-Za-z0-9_]*\s*\([^)]*\)\s*;'),
        'global_variable_pattern': re.compile(r'\b(?:extern\s+)?[A-Za-z_][A-Za-z0-9_]*\s+\**\b[A-Za-z_][A-Za-z0-9_]*\s*(?:=\s*[^;]+)?\s*;'),
        'macro_pattern': re.compile(r'#define\s+\w+\s+.+'),
        'include_pattern': re.compile(r'#include\s*[<"]\S*[>"]')
    }

    # Extract data for each pattern
    data = {name: pattern.findall(content) for name, pattern in patterns.items()}

    return data

def append_data(sheet, data, type_label):
    if data:
        # Add data to the sheet
        for i, item in enumerate(data, start=1):
            print(f"Appending {type_label}: ID={i}, {item}")
            sheet.append([i, item])
    else:
        print(f"No data found for {type_label}. Appending 'not mentioned in the header file'.")
        sheet.append([1, 'not mentioned in the header file'])

def create_excel(out_file):
    res = 'n'
    while res == 'n':
        if os.path.exists(out_file):
            res = input(f"{out_file} already exists. [y to remove / n to change the file name]: ")
            if res == 'y':
                os.remove(out_file)
                print(f"{out_file} has been deleted and a new one will be created.")
            elif res == 'n':
                out_file = input("Enter the new file name: ") + ".xlsx"
                continue
            else:
                print("Invalid input")
                continue
        else:
            print(f"{out_file} does not exist and your file will be created.")
            break

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Create the main sheet for functions
    functions_sheet = workbook.active
    functions_sheet.title = "Functions"
    functions_sheet.append(["ID", "Function Prototype"])

    # Create sheets for other patterns
    sheets = {
        'Global Variables': workbook.create_sheet(title="Global Variables"),
        'Macros': workbook.create_sheet(title="Macros"),
        'Includes': workbook.create_sheet(title="Includes")
    }

    # Adjust the column width
    for sheet in [functions_sheet] + list(sheets.values()):
        sheet.column_dimensions['A'].width = 10
        set
        sheet.column_dimensions['B'].width = 70

    return functions_sheet, sheets, workbook

def set_column_color(sheet, column_letter, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in sheet[column_letter]:
        cell.fill = fill

def header_path():
    while True:
        header = input("Enter the header path: ").strip()
        if os.path.exists(header):
            if os.path.isfile(header) and header.endswith('.h'):
                print("File exists and is a valid header file.")
                break
            else:
                print("The file does not have a .h extension or is not a file.")
        else:
            print("File does not exist or path is invalid.")
    return header

def main():
    header = header_path()
    out_file = input("Enter the output file name (without extension): ").strip() + ".xlsx"
    
    functions_sheet, sheets, workbook = create_excel(out_file)

    # Parse the header file
    data = parse_hf(header)
    
    # Append data to the respective sheets
    append_data(functions_sheet, data['function_pattern'], "Function Prototype")
    append_data(sheets['Global Variables'], data['global_variable_pattern'], "Global Variable Declaration")
    append_data(sheets['Macros'], data['macro_pattern'], "Macro Definition")
    append_data(sheets['Includes'], data['include_pattern'], "Include Statement")
    

    # Set the color of the first column (ID) to yellow for all sheets
    set_column_color(functions_sheet, 'A', 'FFFF00')  # Yellow
    set_column_color(sheets['Global Variables'], 'A', 'FFFF00')  # Yellow
    set_column_color(sheets['Macros'], 'A', 'FFFF00')  # Yellow
    set_column_color(sheets['Includes'], 'A', 'FFFF00')  # Yellow 

    # Set the color of the first column (ID) to yellow for all sheets
    set_column_color(functions_sheet, 'B', 'ADD8E6')  # Yellow
    set_column_color(sheets['Global Variables'], 'B', 'ADD8E6')  # Yellow
    set_column_color(sheets['Macros'], 'B', 'ADD8E6')  # Yellow
    set_column_color(sheets['Includes'], 'B', 'ADD8E6')  # Yellow

    workbook.save(out_file)  # Save the workbook with the user-provided file name
    print("Done")

if __name__ == "__main__":
    main()
